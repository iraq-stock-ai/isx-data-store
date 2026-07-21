import argparse
import io
import os
import re
import sys
import json
from datetime import datetime, timezone

import requests
from bs4 import BeautifulSoup
import openpyxl

LIST_URL = "http://www.isx-iq.net/isxportal/portal/uploadedFilesList.html"
BASE_URL = "http://www.isx-iq.net"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

MIN_EXPECTED_RECORDS = 0
MAX_EXPECTED_RECORDS = 200

FOREIGN_SECTION_MARKERS = ["غير العراقيين", "غيرالعراقيين", "الأجانب", "اجانب", "أجانب"]
BUY_MARKERS = ["المشتراة", "مشتراة"]
SELL_MARKERS = ["المباعة", "مباعة"]
IGNORE_WORDS = {"ISX", "OTC", "TOTAL", "DATE", "TYPE", "BUY", "SELL", "المجموع", "مجموع"}


class QualityGateError(Exception):
    """يُرفع عند فشل أي بوابة تحقق — يوقف السكريبت فوراً دون أي كتابة."""
    pass


def clean_text(txt) -> str:
    if txt is None:
        return ""
    txt = str(txt)
    txt = txt.replace("\xa0", " ").replace("\u200f", "").replace("\u200e", "").replace("\ufeff", "")
    txt = txt.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def parse_date_ddmmyyyy(date_str: str):
    try:
        return datetime.strptime(date_str.strip(), "%d/%m/%Y").date()
    except (ValueError, AttributeError):
        return None


def get_today_daily_report():
    """مطابقة حرفية لدالة get_today_daily_report() بسكربت الأسعار الأصلي."""
    print(f"[بوابة 1] جاري فحص: {LIST_URL}")
    resp = requests.get(LIST_URL, headers=HEADERS, timeout=20)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.content, "html.parser")

    table = soup.find("table")
    if table is None:
        raise QualityGateError("لم يُعثر على جدول الأرشيف بالصفحة الرئيسية.")

    for row in table.find_all("tr"):
        row_text = clean_text(row.get_text())
        if "يومي" not in row_text or "التقرير اليومي" not in row_text:
            continue

        link_tag = row.find("a", href=True)
        if not link_tag:
            continue
        href = link_tag["href"]
        if ".xlsx" not in href.lower() and ".xls" not in href.lower():
            continue

        date_match = re.search(r"(\d{2}/\d{2}/\d{4})", row_text)
        session_date = date_match.group(1) if date_match else None

        full_url = href if href.startswith("http") else BASE_URL + href
        print(f"[بوابة 1] ✅ نجحت — تقرير يومي بتاريخ {session_date}: {full_url}")
        return {"date": session_date, "url": full_url}

    raise QualityGateError("لم يُعثر على أي صف 'التقرير اليومي' بالصفحة الأولى من الأرشيف.")


def validate_report_date(report_date_str: str):
    """مطابقة حرفية لدالة validate_report_date() بسكربت الأسعار الأصلي."""
    report_date = parse_date_ddmmyyyy(report_date_str)
    if report_date is None:
        raise QualityGateError(f"تعذّر تحليل تاريخ التقرير: '{report_date_str}'")

    today = datetime.now(timezone.utc).date()
    diff_days = (today - report_date).days

    if diff_days < 0:
        raise QualityGateError(f"تاريخ التقرير بالمستقبل؟! ({report_date}) — مشبوه، يوقَف التنفيذ.")
    if diff_days > 4:
        raise QualityGateError(f"تاريخ التقرير قديم جداً ({report_date})")

    print(f"[بوابة 2] ✅ نجحت — تاريخ التقرير ({report_date}) ضمن نطاق معقول.")
    return report_date


def download_excel(url: str):
    """مطابقة حرفية لدالة download_excel() بسكربت الأسعار الأصلي."""
    print(f"[بوابة 3] جاري تحميل: {url}")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        raise QualityGateError(f"فشل تحميل ملف Excel: {e}")
    print(f"[بوابة 3] ✅ نجحت — تم تحميل {len(resp.content)} بايت.")
    return resp.content


def extract_session_number(sheet) -> str:
    pattern = re.compile(r"الجلسة\s*[\(\)]?\s*(\d+)")
    for row in sheet.iter_rows(max_row=10, values_only=True):
        for cell in row:
            if cell:
                match = pattern.search(clean_text(cell))
                if match:
                    return match.group(1)
    return "0"


def extract_market_label(section_title_text: str, default_market: str = "النظامي") -> str:
    """تعيين نوع السوق بدقة عالية عبر الكلمات المفتاحية."""
    txt = clean_text(section_title_text)

    if any(k in txt for k in ["غير المفصحة", "غير مفصحة", "الثالث", "otc"]):
        return "الشركات غير المفصحة"
    if any(k in txt for k in ["الثاني", "الثانية"]):
        return "الثاني"
    if any(k in txt for k in ["النظامي", "النظامية"]):
        return "النظامي"

    match = re.search(r"(?:في|منصة|سوق)\s+(.+?)(?:\s+لجلسة|\s*$)", txt)
    if match:
        raw_market = clean_text(match.group(1))
        if "ثاني" in raw_market:
            return "الثاني"
        if "مفصح" in raw_market or "ثالث" in raw_market:
            return "الشركات غير المفصحة"
        if "نظام" in raw_market:
            return "النظامي"

    return default_market


def find_foreign_trading_blocks(wb) -> list:
    blocks = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        num_rows = len(rows)

        current_market = "النظامي"

        for row_idx in range(num_rows):
            curr_text = clean_text(" ".join(str(c) for c in rows[row_idx] if c is not None))

            if "السوق الثاني" in curr_text or "سوق الثاني" in curr_text:
                current_market = "الثاني"
            elif "غير المفصحة" in curr_text or "غير مفصحة" in curr_text:
                current_market = "الشركات غير المفصحة"
            elif "السوق النظامي" in curr_text or "سوق النظامي" in curr_text:
                current_market = "النظامي"

            has_foreign = any(m in curr_text for m in FOREIGN_SECTION_MARKERS)
            has_direction = any(m in curr_text for m in BUY_MARKERS + SELL_MARKERS)

            if has_foreign and has_direction:
                header_lines = []
                start_look = max(0, row_idx - 8)
                end_look = min(row_idx + 3, num_rows)

                for w_idx in range(start_look, end_look):
                    w_text = clean_text(" ".join(str(c) for c in rows[w_idx] if c is not None))
                    header_lines.append(w_text)

                full_header = " ".join(header_lines)
                market_label = extract_market_label(full_header, default_market=current_market)

                blocks.append((sheet_name, row_idx, full_header, market_label))

    return blocks


def parse_foreign_section(sheet, start_row_idx: int, market_label: str, direction: str) -> list:
    records = []
    symbol_pattern = re.compile(r"\b([A-Z]{3,6})\b")
    rows = list(sheet.iter_rows(values_only=True))
    MAX_ROWS = 60

    for offset in range(1, MAX_ROWS):
        idx = start_row_idx + offset
        if idx >= len(rows):
            break

        row = rows[idx]
        row_cells = [clean_text(c) for c in row]
        row_text = " ".join(c for c in row_cells if c)

        if not row_text:
            continue

        if any(term in row_text for term in ["المجموع الكلي", "مجموع الكلي"]):
            break
        if any(m in row_text for m in FOREIGN_SECTION_MARKERS) and any(m in row_text for m in BUY_MARKERS + SELL_MARKERS):
            break

        if "مجموع" in row_text or "قطاع" in row_text:
            continue

        symbol = None
        for cell in row_cells:
            cell_clean = cell.strip().upper()
            m = symbol_pattern.search(cell_clean)
            if m:
                cand = m.group(1)
                if cand not in IGNORE_WORDS:
                    symbol = cand
                    break

        if not symbol:
            continue

        numbers = []
        for cell in row_cells:
            cell_num = cell.replace(",", "").strip()
            if re.match(r"^-?\d+(\.\d+)?$", cell_num):
                numbers.append(cell_num)

        if len(numbers) < 3:
            continue

        trades_raw, shares_raw, value_raw = numbers[-3], numbers[-2], numbers[-1]
        try:
            trades = int(float(trades_raw))
            shares = int(float(shares_raw))
            value = int(float(value_raw))
        except ValueError:
            continue

        records.append({
            "symbol": symbol,
            "market": market_label,
            "direction": direction,
            "trades": trades,
            "shares": shares,
            "value": value,
        })

    return records


def parse_daily_excel(excel_bytes: bytes) -> dict:
    """يستخرج بيانات "غير العراقيين" وتصيغ القاموس النهائي.
    اسم الدالة مطابق عمداً لسكربت الأسعار الأصلي (parse_daily_excel)
    للحفاظ على نفس البنية العامة، رغم اختلاف المحتوى المُستخرَج."""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    session_number = "0"
    if wb.sheetnames:
        session_number = extract_session_number(wb[wb.sheetnames[0]])

    blocks = find_foreign_trading_blocks(wb)
    if not blocks:
        return {"session_number": session_number, "records": []}

    all_records = []
    for sheet_name, row_idx, row_text, market_label in blocks:
        sheet = wb[sheet_name]

        direction = "buy" if any(m in row_text for m in BUY_MARKERS) else (
            "sell" if any(m in row_text for m in SELL_MARKERS) else None
        )
        if direction is None:
            continue

        section_records = parse_foreign_section(sheet, row_idx, market_label, direction)
        all_records.extend(section_records)

    return {"session_number": session_number, "records": all_records}


def check_records_count(parsed: dict):
    """مطابقة لروح دالة check_symbol_count() بسكربت الأسعار الأصلي،
    لكن بحدود مختلفة تناسب طبيعة بيانات غير العراقيين (قد تكون صفراً)."""
    count = len(parsed["records"])
    if count > MAX_EXPECTED_RECORDS:
        raise QualityGateError(
            f"عدد سجلات غير العراقيين المستخرجة ({count}) يتجاوز الحد الأقصى المعقول ({MAX_EXPECTED_RECORDS})."
        )
    if count == 0:
        print("[بوابة 4] ⚠️ لا توجد أي سجلات حركة غير عراقيين بهذا اليوم — نتيجة محتملة وطبيعية، وليست خطأ.")
    else:
        print(f"[بوابة 4] ✅ نجحت — {count} سجل حركة غير عراقيين مستخرَج.")


def validate_records(parsed: dict) -> list:
    """فحص جودة السجلات — مطابقة لروح دالة validate_records() الأصلية،
    لكن بمعايير تناسب طبيعة هذه البيانات (لا أسعار Open/High/Low/Close هنا)."""
    valid = []
    rejected = []

    for r in parsed["records"]:
        problems = []
        if r["trades"] < 0 or r["shares"] < 0 or r["value"] < 0:
            problems.append("قيمة سالبة غير منطقية (صفقات/أسهم/قيمة)")
        if r["shares"] == 0 and r["value"] != 0:
            problems.append("عدد أسهم صفر لكن قيمة غير صفرية — تناقض")

        if problems:
            rejected.append(f"{r['symbol']}: " + " | ".join(problems))
        else:
            valid.append(r)

    total = len(parsed["records"])
    ratio = len(rejected) / total if total else 0.0
    print(f"[بوابة 5] فحص الجودة: {len(valid)} سليم، {len(rejected)} مرفوض.")

    if ratio > 0.10 and total > 0:
        raise QualityGateError(f"نسبة السجلات المرفوضة ({ratio:.1%}) تتجاوز الحد المسموح به.")

    print("[بوابة 5+6] ✅ نجحتا — البيانات متناسقة وجاهزة للدمج.")
    return valid


def load_existing(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(data: dict, path: str):
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)


def main():
    parser = argparse.ArgumentParser(description="التحديث اليومي التلقائي لأرشيف تداول غير العراقيين بسوق العراق")
    parser.add_argument("--existing", default="isx_foreign_trading.json")
    parser.add_argument("--output", default="isx_foreign_trading.json")
    args = parser.parse_args()

    try:
        report = get_today_daily_report()
        validate_report_date(report["date"])
        excel_bytes = download_excel(report["url"])
        parsed = parse_daily_excel(excel_bytes)
        check_records_count(parsed)
        valid_records = validate_records(parsed)

    except QualityGateError as e:
        print(f"\n❌ توقف التنفيذ — فشلت بوابة تحقق: {e}", file=sys.stderr)
        sys.exit(1)

    print("\n✅ كل بوابات التحقق نجحت. جاري الدمج مع ملف الـ JSON الحالي...")

    data = load_existing(args.existing)

    added_count = 0
    for rec in valid_records:
        symbol = rec["symbol"]
        if symbol not in data:
            data[symbol] = []

        # منع التكرار: لو نفس (date + direction) موجود مسبقاً لنفس
        # الرمز، يُتجاهل — يمنع تكرار الإدخال لو اشتغل السكربت عدة
        # مرات بنفس اليوم (2، 4، 6، 8 مساءً) ووجد نفس البيانات.
        duplicate = any(
            r.get("date") == report["date"] and r.get("direction") == rec["direction"]
            for r in data[symbol]
        )
        if duplicate:
            continue

        data[symbol].insert(0, {
            "date": report["date"],
            "sessionNumber": parsed.get("session_number"),
            "market": rec["market"],
            "direction": rec["direction"],
            "trades": rec["trades"],
            "shares": rec["shares"],
            "value": rec["value"],
        })
        added_count += 1

    if added_count == 0:
        print("لا توجد سجلات جديدة لإضافتها (اليوم موجود مسبقاً من محاولة سابقة). الملف يبقى كما هو دون تعديل.")
        sys.exit(0)

    save_json(data, args.output)
    print(f"\n✅ تم التحديث بنجاح! أُضيف {added_count} سجل جديد ليوم {report['date']}.")
    sys.exit(0)


if __name__ == "__main__":
    main()
