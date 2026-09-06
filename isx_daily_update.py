import argparse
import io
import os
import re
import sys
import json
from datetime import datetime, timezone, date, timedelta
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
import openpyxl

LIST_URL = "http://www.isx-iq.net/isxportal/portal/uploadedFilesList.html"
BASE_URL = "http://www.isx-iq.net"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

# عدد الصفحات القصوى التي يمكن فحصها للعثور على التقرير اليومي
MAX_ARCHIVE_PAGES = 10

# حدود بوابة التحقق
MIN_EXPECTED_SYMBOLS = 30
MAX_EXPECTED_SYMBOLS = 130
MAX_REJECTED_RATIO = 0.10  # 10%


class QualityGateError(Exception):
    """يُرفع عند فشل أي بوابة تحقق — يوقف السكريبت فوراً دون أي كتابة."""
    pass


def clean_text(txt) -> str:
    if txt is None:
        return ""
    txt = str(txt).replace("\r", " ").replace("\n", " ").replace("\t", " ")
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def parse_date_ddmmyyyy(date_str: str):
    try:
        return datetime.strptime(date_str.strip(), "%d/%m/%Y").date()
    except (ValueError, AttributeError):
        return None


def extract_report_from_page(soup, page_number):
    """
    البحث عن التقرير اليومي داخل صفحة واحدة من أرشيف ISX.

    لا نعتمد على أول جدول فقط، لأن الصفحة قد تحتوي على أكثر من جدول
    وقد يكون التقرير اليومي موجوداً في جدول آخر.
    """

    # نفحص جميع الصفوف الموجودة في جميع الجداول
    for table in soup.find_all("table"):
        for row in table.find_all("tr"):

            row_text = clean_text(row.get_text(" ", strip=True))

            # التقرير المطلوب يجب أن يكون تقريراً يومياً
            if "التقرير اليومي" not in row_text and "تقرير يومي" not in row_text:
                continue

            # نستبعد التقارير الشهرية صراحةً
            if "التقرير الشهري" in row_text or "تقرير شهري" in row_text:
                continue

            # نبحث عن رابط Excel داخل نفس الصف
            excel_link = None

            for link_tag in row.find_all("a", href=True):
                href = clean_text(link_tag.get("href"))

                if not href:
                    continue

                href_lower = href.lower()

                if (
                    ".xlsx" in href_lower
                    or ".xls" in href_lower
                    or "excel" in href_lower
                ):
                    excel_link = link_tag
                    break

            if excel_link is None:
                continue

            href = clean_text(excel_link.get("href"))

            # استخراج التاريخ من الصف
            date_match = re.search(
                r"(\d{2}/\d{2}/\d{4})",
                row_text
            )

            session_date = date_match.group(1) if date_match else None

            # بناء الرابط بشكل آمن سواء كان:
            # /path/file.xls
            # أو path/file.xls
            # أو رابطاً كاملاً
            full_url = urljoin(BASE_URL + "/", href)

            print(
                f"[بوابة 1] 🔎 تم العثور على تقرير يومي "
                f"في الصفحة {page_number}"
            )

            print(
                f"[بوابة 1] التاريخ: {session_date}"
            )

            print(
                f"[بوابة 1] الرابط: {full_url}"
            )

            return {
                "date": session_date,
                "url": full_url,
                "page": page_number
            }

    return None


def get_today_daily_report():
    """
    البحث عن أحدث تقرير يومي في أرشيف ISX.

    الصفحة الأولى قد تكون مسيطراً عليها بالتقارير الشهرية،
    لذلك نفحص الصفحات بالتتابع حتى العثور على التقرير اليومي.

    يتم فحص عدد محدود من الصفحات لمنع الدوران غير المنتهي.
    """

    print(
        f"[بوابة 1] بدء البحث عن التقرير اليومي في أرشيف ISX"
    )

    session = requests.Session()
    session.headers.update(HEADERS)

    # سنجمع التقارير اليومية التي نعثر عليها
    candidates = []

    for page_number in range(1, MAX_ARCHIVE_PAGES + 1):

        # الصفحة الأولى بدون parameter
        if page_number == 1:
            page_url = LIST_URL
        else:
            # محاولة نمط pagination المعتاد
            page_url = f"{LIST_URL}?page={page_number}"

        print(
            f"[بوابة 1] جاري فحص الصفحة {page_number}: {page_url}"
        )

        try:
            resp = session.get(
                page_url,
                timeout=20
            )
            resp.raise_for_status()

        except requests.RequestException as e:

            print(
                f"[بوابة 1] ⚠️ تعذر تحميل الصفحة {page_number}: {e}"
            )

            # لا نوقف البرنامج فوراً إذا كانت صفحة لاحقة غير متاحة
            # بل ننتقل للصفحة التالية.
            continue

        soup = BeautifulSoup(
            resp.content,
            "html.parser"
        )

        result = extract_report_from_page(
            soup,
            page_number
        )

        if result:
            candidates.append(result)

            # بما أننا نبحث من الصفحة الأولى إلى الأبعد،
            # أول تقرير يومي هو عادةً الأحدث.
            #
            # لكننا لا نعود مباشرةً قبل التحقق من أن التاريخ صالح.
            report_date = parse_date_ddmmyyyy(
                result.get("date")
            )

            if report_date is not None:

                today = datetime.now(
                    timezone.utc
                ).date()

                diff_days = (
                    today - report_date
                ).days

                # تقرير حديث ومعقول
                if 0 <= diff_days <= 4:

                    print(
                        f"[بوابة 1] ✅ نجحت — "
                        f"تم العثور على تقرير يومي حديث "
                        f"في الصفحة {page_number}: "
                        f"{result.get('date')}"
                    )

                    return result

                # إذا كان التاريخ غير حديث، نكمل البحث
                print(
                    f"[بوابة 1] ⚠️ التقرير الموجود في الصفحة "
                    f"{page_number} تاريخه {result.get('date')} "
                    f"وليس ضمن النطاق الحديث، نتابع البحث."
                )

            else:
                print(
                    f"[بوابة 1] ⚠️ تم العثور على تقرير يومي "
                    f"لكن تعذر استخراج التاريخ، نتابع البحث."
                )

    # إذا لم نجد تقريراً حديثاً لكن وجدنا تقريراً يومياً
    # نعيد أقرب مرشح حتى تتولى بوابة التاريخ التحقق النهائي.
    if candidates:

        print(
            f"[بوابة 1] ⚠️ تم العثور على {len(candidates)} "
            f"تقرير/تقارير يومية، لكن لا يوجد تقرير حديث مطابق "
            f"للنطاق المتوقع."
        )

        # اختيار التقرير الذي يملك تاريخاً صالحاً والأقرب لليوم
        today = datetime.now(
            timezone.utc
        ).date()

        dated_candidates = []

        for candidate in candidates:

            report_date = parse_date_ddmmyyyy(
                candidate.get("date")
            )

            if report_date is not None:
                dated_candidates.append(
                    (
                        abs((today - report_date).days),
                        candidate
                    )
                )

        if dated_candidates:

            dated_candidates.sort(
                key=lambda x: x[0]
            )

            selected = dated_candidates[0][1]

            print(
                f"[بوابة 1] سيتم استخدام أقرب تقرير يومي "
                f"بتاريخ {selected.get('date')}"
            )

            return selected

        # يوجد تقرير يومي لكن بدون تاريخ
        return candidates[0]

    raise QualityGateError(
        f"لم يُعثر على أي تقرير يومي ضمن أول "
        f"{MAX_ARCHIVE_PAGES} صفحات من الأرشيف."
    )


def validate_report_date(report_date_str: str):
    report_date = parse_date_ddmmyyyy(report_date_str)

    if report_date is None:
        raise QualityGateError(
            f"تعذّر تحليل تاريخ التقرير: '{report_date_str}'"
        )

    today = datetime.now(
        timezone.utc
    ).date()

    diff_days = (
        today - report_date
    ).days

    if diff_days < 0:
        raise QualityGateError(
            f"تاريخ التقرير بالمستقبل؟! ({report_date}) "
            f"— مشبوه، يوقَف التنفيذ."
        )

    if diff_days > 4:
        raise QualityGateError(
            f"تاريخ التقرير قديم جداً ({report_date})"
        )

    print(
        f"[بوابة 2] ✅ نجحت — "
        f"تاريخ التقرير ({report_date}) ضمن نطاق معقول."
    )

    return report_date


def download_excel(url: str):
    print(f"[بوابة 3] جاري تحميل: {url}")

    try:
        resp = requests.get(
            url,
            headers=HEADERS,
            timeout=30
        )

        resp.raise_for_status()

    except Exception as e:
        raise QualityGateError(
            f"فشل تحميل ملف Excel: {e}"
        )

    if not resp.content:
        raise QualityGateError(
            "تم تحميل ملف Excel فارغ."
        )

    print(
        f"[بوابة 3] ✅ نجحت — "
        f"تم تحميل {len(resp.content)} بايت."
    )

    return resp.content


def find_header_row_and_map(sheet):
    """
    البحث بالأسماء العربية المتوقعة من ملف الإكسل
    وربطها بالمفاتيح الإنجليزية للـ JSON التاريخي.
    """

    HEADER_ALIASES = {
        "open": [
            "افتتاح",
            "فتح",
            "سعر الفتح"
        ],

        "high": [
            "اعلى سعر",
            "أعلى سعر",
            "أعلى",
            "اعلى"
        ],

        "low": [
            "ادنى سعر",
            "أدنى سعر",
            "أدنى",
            "ادنى"
        ],

        "close": [
            "سعر الاغلاق",
            "سعر الإغلاق",
            "اغلاق",
            "إغلاق",
            "سعر القفل"
        ],

        "volume": [
            "الاسهم المتداولة",
            "الأسهم المتداولة",
            "حجم التداول"
        ],

        "value": [
            "القيمة المتداولة",
            "القيمة المتدوالة",
            "قيمة التداول"
        ],

        "trades": [
            "الصفقات",
            "عدد الصفقات"
        ],
    }

    SYMBOL_HEADER_ALIASES = [
        "رمز الشركة",
        "الرمز",
        "رمز",
        "Symbol"
    ]

    for row_idx, row in enumerate(
        sheet.iter_rows(
            max_row=15,
            values_only=True
        )
    ):

        row_texts = [
            clean_text(c)
            for c in row
        ]

        col_map = {}
        symbol_col = None

        for col_idx, text in enumerate(row_texts):

            if not text:
                continue

            for field, aliases in HEADER_ALIASES.items():

                if text in aliases:
                    col_map[field] = col_idx

            if text in SYMBOL_HEADER_ALIASES:
                symbol_col = col_idx

        # التحقق من الأعمدة المالية الأساسية
        core_fields = {
            "open",
            "high",
            "low",
            "close"
        }

        if core_fields.issubset(
            col_map.keys()
        ):
            return (
                row_idx,
                col_map,
                symbol_col
            )

    return None, {}, None


def extract_session_date_from_excel(sheet) -> str:
    date_pattern = re.compile(
        r"(\d{4}/\d{1,2}/\d{1,2})"
    )

    for row in sheet.iter_rows(
        max_row=5,
        values_only=True
    ):

        for cell in row:

            if cell:

                match = date_pattern.search(
                    str(cell)
                )

                if match:

                    y, m, d = match.group(1).split("/")

                    return (
                        f"{int(d):02d}/"
                        f"{int(m):02d}/"
                        f"{y}"
                    )

    return None


def parse_daily_excel(excel_bytes: bytes) -> dict:
    """
    تستخرج البيانات وتصيغ القاموس النهائي
    بالمفاتيح الإنجليزية المتوافقة مع
    أرشيف الـ JSON التاريخي.
    """

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(excel_bytes),
            data_only=True
        )

    except Exception as e:
        raise QualityGateError(
            f"تعذر فتح ملف Excel: {e}"
        )

    results = {}

    pure_symbol_pattern = re.compile(
        r"^[A-Z]{3,5}$"
    )

    bracket_pattern = re.compile(
        r"\((_?[A-Z]{3,5})\)"
    )

    session_date = None

    if wb.sheetnames:

        session_date = (
            extract_session_date_from_excel(
                wb[wb.sheetnames[0]]
            )
        )

    for sheet_name in wb.sheetnames:

        sheet = wb[sheet_name]

        header_row_idx, col_map, symbol_col = (
            find_header_row_and_map(sheet)
        )

        if header_row_idx is None:
            continue

        for row_idx, row in enumerate(
            sheet.iter_rows(
                values_only=True
            )
        ):

            if row_idx <= header_row_idx:
                continue

            row_str = [
                clean_text(c)
                for c in row
            ]

            symbol = ""

            if (
                symbol_col is not None
                and symbol_col < len(row_str)
            ):

                candidate = row_str[symbol_col]

                if pure_symbol_pattern.match(
                    candidate
                ):

                    symbol = candidate

            if not symbol:

                for text in row_str:

                    if not text:
                        continue

                    if (
                        pure_symbol_pattern.match(text)
                        and text not in [
                            "ISX",
                            "OTC",
                            "NONE"
                        ]
                    ):

                        symbol = text
                        break

                    m = bracket_pattern.search(
                        text
                    )

                    if m:

                        symbol = m.group(1)
                        break

            if not symbol:
                continue

            if symbol in results:
                continue

            def get_field(field_name):

                col = col_map.get(
                    field_name
                )

                if (
                    col is None
                    or col >= len(row_str)
                ):
                    return "-"

                v = row_str[col]

                return (
                    v
                    if v and v != "None"
                    else "-"
                )

            results[symbol] = {
                "date": session_date or "",
                "open": get_field("open"),
                "high": get_field("high"),
                "low": get_field("low"),
                "close": get_field("close"),
                "volume": get_field("volume"),
                "value": get_field("value"),
                "trades": get_field("trades"),
            }

    return results


def check_symbol_count(day_data: dict):

    count = len(day_data)

    if (
        count < MIN_EXPECTED_SYMBOLS
        or count > MAX_EXPECTED_SYMBOLS
    ):

        raise QualityGateError(
            f"عدد الأسهم المستخرجة ({count}) "
            f"خارج النطاق المتوقع "
            f"({MIN_EXPECTED_SYMBOLS}-{MAX_EXPECTED_SYMBOLS})."
        )

    print(
        f"[بوابة 4] ✅ نجحت — "
        f"تم استخراج {count} سهم "
        f"وتوزيع مسمياتها الإنجليزية."
    )


def validate_records(day_data: dict):
    """
    فحص جودة السجلات بناءً على
    التسميات الجديدة المتوافقة.
    """

    valid = {}
    rejected = []

    for symbol, r in day_data.items():

        def as_float(key):

            v = r.get(key)

            if v in (
                None,
                "-",
                ""
            ):
                return None

            try:
                return float(
                    str(v).replace(",", "")
                )

            except (
                ValueError,
                TypeError
            ):
                return None

        high = as_float("high")
        low = as_float("low")
        close = as_float("close")

        problems = []

        if (
            high is not None
            and low is not None
            and high < low
        ):

            problems.append(
                f"High({high}) < Low({low})"
            )

        if (
            close is not None
            and high is not None
            and low is not None
            and not (
                low <= close <= high
            )
        ):

            problems.append(
                f"Close({close}) "
                f"خارج النطاق [{low}, {high}]"
            )

        if problems:

            rejected.append(
                f"{symbol}: "
                + " | ".join(problems)
            )

        else:

            valid[symbol] = r

    ratio = (
        len(rejected) / len(day_data)
        if day_data
        else 1.0
    )

    print(
        f"[بوابة 5] فحص الجودة: "
        f"{len(valid)} سليم، "
        f"{len(rejected)} مرفوض."
    )

    if ratio > MAX_REJECTED_RATIO:

        raise QualityGateError(
            f"نسبة السجلات المرفوضة "
            f"({ratio:.1%}) "
            f"تتجاوز الحد المسموح به."
        )

    print(
        "[بوابة 5+6] ✅ نجحتا — "
        "البيانات متناسقة وجاهزة للدمج."
    )

    return valid


def load_existing(path: str) -> dict:

    if not os.path.exists(path):
        return {}

    try:

        with open(
            path,
            "r",
            encoding="utf-8"
        ) as f:

            return json.load(f)

    except Exception as e:

        raise QualityGateError(
            f"تعذر قراءة ملف JSON الحالي: {e}"
        )


def save_json(data: dict, path: str):

    tmp_path = path + ".tmp"

    with open(
        tmp_path,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            data,
            f,
            ensure_ascii=False,
            indent=2
        )

    os.replace(
        tmp_path,
        path
    )


def main():

    parser = argparse.ArgumentParser(
        description=(
            "التحديث اليومي التلقائي "
            "لأرشيف أسهم ISX"
        )
    )

    parser.add_argument(
        "--existing",
        default="isx_history_all.json"
    )

    parser.add_argument(
        "--output",
        default="isx_history_all.json"
    )

    args = parser.parse_args()

    try:

        # بوابة 1
        report = get_today_daily_report()

        # بوابة 2
        validate_report_date(
            report["date"]
        )

        # بوابة 3
        excel_bytes = download_excel(
            report["url"]
        )

        # تحليل Excel
        day_data = parse_daily_excel(
            excel_bytes
        )

        # بوابة 4
        check_symbol_count(
            day_data
        )

        # بوابات 5 + 6
        valid_data = validate_records(
            day_data
        )

    except QualityGateError as e:

        print(
            f"\n❌ توقف التنفيذ — "
            f"فشلت بوابة تحقق: {e}",
            file=sys.stderr
        )

        sys.exit(1)

    print(
        "\n✅ كل بوابات التحقق نجحت "
        "وبنيت الهياكل بنجاح. "
        "جاري الدمج مع ملف الـ JSON التاريخي..."
    )

    data = load_existing(
        args.existing
    )

    existing_dates = {
        sym: {
            r["date"]
            for r in recs
            if r.get("date")
        }
        for sym, recs in data.items()
    }

    added_count = 0

    for symbol, record in valid_data.items():

        record_date = (
            record.get("date")
            or report["date"]
        )

        record["date"] = record_date

        if symbol not in data:

            data[symbol] = []
            existing_dates[symbol] = set()

        if (
            record_date
            in existing_dates[symbol]
        ):
            continue

        data[symbol].insert(
            0,
            record
        )

        existing_dates[symbol].add(
            record_date
        )

        added_count += 1

    if added_count == 0:

        print(
            "لا توجد سجلات جديدة لإضافتها. "
            "الملف يبقى كما هو دون تعديل."
        )

        sys.exit(0)

    save_json(
        data,
        args.output
    )

    print(
        "\n✅ تم التحديث بنجاح! "
        "حُفظت البيانات بالمسميات الإنجليزية "
        "لتتطابق تماماً وبشكل سلس "
        "مع مستودع GitHub."
    )

    print(
        f"📊 عدد السجلات الجديدة المضافة: "
        f"{added_count}"
    )

    sys.exit(0)


if __name__ == "__main__":
    main()
